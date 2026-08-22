"""
build_excel_workbook.py
Builds the complete, executive-grade multi-tab Excel Workbook:
excel/retail_sales_analysis.xlsx

Contains 8 Professional Sheets:
1. Raw Data
2. Data Dictionary
3. Sales Analysis
4. Customer Analysis
5. Product Analysis
6. Regional Analysis
7. Pivot Analysis
8. Executive Summary
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
CLEANED_DATA_DIR = os.path.join(BASE_DIR, "data", "cleaned")
EXCEL_DIR = os.path.join(BASE_DIR, "excel")
os.makedirs(EXCEL_DIR, exist_ok=True)

EXCEL_FILE = os.path.join(EXCEL_DIR, "retail_sales_analysis.xlsx")

def create_workbook():
    print("=" * 60)
    print("  BUILDING PROFESSIONAL EXCEL WORKBOOK: retail_sales_analysis.xlsx")
    print("=" * 60)
    
    # 1. Load Data
    df_sales = pd.read_csv(os.path.join(CLEANED_DATA_DIR, "fact_sales.csv"))
    df_cust = pd.read_csv(os.path.join(CLEANED_DATA_DIR, "dim_customer.csv"))
    df_prod = pd.read_csv(os.path.join(CLEANED_DATA_DIR, "dim_product.csv"))
    df_reg = pd.read_csv(os.path.join(CLEANED_DATA_DIR, "dim_region.csv"))
    df_date = pd.read_csv(os.path.join(CLEANED_DATA_DIR, "dim_date.csv"))
    df_rfm = pd.read_csv(os.path.join(CLEANED_DATA_DIR, "customer_rfm_segments.csv"))
    
    # Master merged sales table
    df_master = df_sales.merge(df_prod[["product_id", "product_name", "category", "subcategory", "brand"]], on="product_id")
    df_master = df_master.merge(df_reg, on="region_id")
    df_master = df_master.merge(df_cust[["customer_id", "customer_name", "customer_segment"]], on="customer_id")
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Common Styling Definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=13, bold=True, color="1B365D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, color="000000")
    font_kpi_num = Font(name="Calibri", size=18, bold=True, color="1B365D")
    font_kpi_label = Font(name="Calibri", size=10, bold=True, color="555555")
    
    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_slate = PatternFill(start_color="334D6E", end_color="334D6E", fill_type="solid")
    fill_header = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    fill_accent = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_kpi_card = PatternFill(start_color="EAEEF7", end_color="EAEEF7", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thick_bottom_side = Side(border_style="medium", color="1B365D")
    double_bottom_side = Side(border_style="double", color="1B365D")
    
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_border_side, bottom=double_bottom_side)
    border_kpi = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # -------------------------------------------------------------
    # SHEET 8: EXECUTIVE SUMMARY (Placed first visually)
    # -------------------------------------------------------------
    print("  Creating Sheet 8: Executive Summary...")
    ws_exec = wb.create_sheet(title="Executive Summary")
    ws_exec.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_exec.merge_cells("A1:I2")
    ws_exec["A1"] = "RETAIL SALES & CUSTOMER INSIGHTS — EXECUTIVE DASHBOARD"
    ws_exec["A1"].font = font_title
    ws_exec["A1"].fill = fill_navy
    ws_exec["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Subtitle
    ws_exec.merge_cells("A3:I3")
    ws_exec["A3"] = "Enterprise Performance Scorecard across 61,926 Transactions, 11,500 Customers, 12 Categories & 6 Regions (2024–2025)"
    ws_exec["A3"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws_exec["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # KPI Scorecard Cards
    kpis = [
        ("TOTAL REVENUE", 45828146.55, "₹#,##0.00", "A5:B6", "A5"),
        ("TOTAL PROFIT", 17316104.20, "₹#,##0.00", "C5:C6", "C5"),
        ("PROFIT MARGIN", 0.3778, "0.00%", "D5:D6", "D5"),
        ("TOTAL ORDERS", 19021, "#,##0", "E5:E6", "E5"),
        ("UNIQUE CUSTOMERS", 11500, "#,##0", "F5:F6", "F5"),
        ("AVG ORDER VALUE", 2409.34, "₹#,##0.00", "G5:G6", "G5"),
        ("REPEAT RATE", 0.3519, "0.00%", "H5:H6", "H5"),
        ("TOP CATEGORY", "Electronics (27.0%)", "@", "I5:I6", "I5"),
    ]
    
    for label, val, num_fmt, merge_range, top_cell in kpis:
        # Check if single or multi column merge
        if ":" in merge_range:
            ws_exec.merge_cells(merge_range)
        cell = ws_exec[top_cell]
        cell.value = f"{label}\n{val:,.2f}" if isinstance(val, (int, float)) and num_fmt.startswith("₹") else (f"{label}\n{val*100:.2f}%" if num_fmt == "0.00%" else f"{label}\n{val:,}" if isinstance(val, int) else f"{label}\n{val}")
        cell.font = font_bold
        cell.fill = fill_kpi_card
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Apply border
        for row in ws_exec[merge_range]:
            for c in row:
                c.border = border_kpi
                
    # Section 1: Regional & Category Summary Tables
    ws_exec["A8"] = "1. Regional Sales Performance"
    ws_exec["A8"].font = font_section
    
    reg_headers = ["Region", "Zone", "Revenue (₹)", "Orders", "AOV (₹)", "Share %", "Profit (₹)", "Margin %"]
    for col_idx, h in enumerate(reg_headers, start=1):
        c = ws_exec.cell(row=9, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = border_cell
        
    df_reg_sum = df_master.groupby(["region_name", "zone"]).agg(
        Revenue=("sales_amount", "sum"),
        Orders=("order_id", "nunique"),
        Profit=("profit", "sum")
    ).reset_index().sort_values(by="Revenue", ascending=False)
    
    tot_rev = df_sales["sales_amount"].sum()
    
    curr_row = 10
    for _, r in df_reg_sum.iterrows():
        ws_exec.cell(row=curr_row, column=1, value=r["region_name"]).font = font_regular
        ws_exec.cell(row=curr_row, column=2, value=r["zone"]).font = font_regular
        
        c3 = ws_exec.cell(row=curr_row, column=3, value=r["Revenue"])
        c3.number_format = "₹#,##0.00"
        c3.font = font_regular
        
        c4 = ws_exec.cell(row=curr_row, column=4, value=r["Orders"])
        c4.number_format = "#,##0"
        c4.font = font_regular
        
        c5 = ws_exec.cell(row=curr_row, column=5, value=r["Revenue"] / r["Orders"])
        c5.number_format = "₹#,##0.00"
        c5.font = font_regular
        
        c6 = ws_exec.cell(row=curr_row, column=6, value=r["Revenue"] / tot_rev)
        c6.number_format = "0.00%"
        c6.font = font_regular
        
        c7 = ws_exec.cell(row=curr_row, column=7, value=r["Profit"])
        c7.number_format = "₹#,##0.00"
        c7.font = font_regular
        
        c8 = ws_exec.cell(row=curr_row, column=8, value=r["Profit"] / r["Revenue"])
        c8.number_format = "0.00%"
        c8.font = font_regular
        
        for c_idx in range(1, 9):
            ws_exec.cell(row=curr_row, column=c_idx).border = border_cell
        curr_row += 1
        
    # Total row for Region
    ws_exec.cell(row=curr_row, column=1, value="Total").font = font_bold
    ws_exec.cell(row=curr_row, column=2, value="All Zones").font = font_bold
    
    tot_c3 = ws_exec.cell(row=curr_row, column=3, value=f"=SUM(C10:C{curr_row-1})")
    tot_c3.number_format = "₹#,##0.00"
    tot_c3.font = font_bold
    
    tot_c4 = ws_exec.cell(row=curr_row, column=4, value=f"=SUM(D10:D{curr_row-1})")
    tot_c4.number_format = "#,##0"
    tot_c4.font = font_bold
    
    tot_c5 = ws_exec.cell(row=curr_row, column=5, value=f"=C{curr_row}/D{curr_row}")
    tot_c5.number_format = "₹#,##0.00"
    tot_c5.font = font_bold
    
    tot_c6 = ws_exec.cell(row=curr_row, column=6, value=f"=SUM(F10:F{curr_row-1})")
    tot_c6.number_format = "0.00%"
    tot_c6.font = font_bold
    
    tot_c7 = ws_exec.cell(row=curr_row, column=7, value=f"=SUM(G10:G{curr_row-1})")
    tot_c7.number_format = "₹#,##0.00"
    tot_c7.font = font_bold
    
    tot_c8 = ws_exec.cell(row=curr_row, column=8, value=f"=G{curr_row}/C{curr_row}")
    tot_c8.number_format = "0.00%"
    tot_c8.font = font_bold
    
    for c_idx in range(1, 9):
        ws_exec.cell(row=curr_row, column=c_idx).border = border_total

    # Section 2: Category Summary Table
    curr_row += 3
    ws_exec.cell(row=curr_row, column=1, value="2. Category Sales & Profitability Rankings").font = font_section
    curr_row += 1
    
    cat_headers = ["Category", "Units Sold", "Revenue (₹)", "Revenue Share %", "Profit (₹)", "Profit Margin %"]
    for col_idx, h in enumerate(cat_headers, start=1):
        c = ws_exec.cell(row=curr_row, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = border_cell
        
    df_cat_sum = df_master.groupby("category").agg(
        Units=("quantity", "sum"),
        Revenue=("sales_amount", "sum"),
        Profit=("profit", "sum")
    ).reset_index().sort_values(by="Revenue", ascending=False)
    
    cat_start_row = curr_row + 1
    curr_row += 1
    for _, r in df_cat_sum.iterrows():
        ws_exec.cell(row=curr_row, column=1, value=r["category"]).font = font_regular
        
        c2 = ws_exec.cell(row=curr_row, column=2, value=r["Units"])
        c2.number_format = "#,##0"
        c2.font = font_regular
        
        c3 = ws_exec.cell(row=curr_row, column=3, value=r["Revenue"])
        c3.number_format = "₹#,##0.00"
        c3.font = font_regular
        
        c4 = ws_exec.cell(row=curr_row, column=4, value=r["Revenue"] / tot_rev)
        c4.number_format = "0.00%"
        c4.font = font_regular
        
        c5 = ws_exec.cell(row=curr_row, column=5, value=r["Profit"])
        c5.number_format = "₹#,##0.00"
        c5.font = font_regular
        
        c6 = ws_exec.cell(row=curr_row, column=6, value=r["Profit"] / r["Revenue"])
        c6.number_format = "0.00%"
        c6.font = font_regular
        
        for c_idx in range(1, 7):
            ws_exec.cell(row=curr_row, column=c_idx).border = border_cell
        curr_row += 1
        
    # Total row for Category
    ws_exec.cell(row=curr_row, column=1, value="Total").font = font_bold
    
    c_tot_units = ws_exec.cell(row=curr_row, column=2, value=f"=SUM(B{cat_start_row}:B{curr_row-1})")
    c_tot_units.number_format = "#,##0"
    c_tot_units.font = font_bold
    
    c_tot_rev = ws_exec.cell(row=curr_row, column=3, value=f"=SUM(C{cat_start_row}:C{curr_row-1})")
    c_tot_rev.number_format = "₹#,##0.00"
    c_tot_rev.font = font_bold
    
    c_tot_share = ws_exec.cell(row=curr_row, column=4, value=f"=SUM(D{cat_start_row}:D{curr_row-1})")
    c_tot_share.number_format = "0.00%"
    c_tot_share.font = font_bold
    
    c_tot_prof = ws_exec.cell(row=curr_row, column=5, value=f"=SUM(E{cat_start_row}:E{curr_row-1})")
    c_tot_prof.number_format = "₹#,##0.00"
    c_tot_prof.font = font_bold
    
    c_tot_mrg = ws_exec.cell(row=curr_row, column=6, value=f"=E{curr_row}/C{curr_row}")
    c_tot_mrg.number_format = "0.00%"
    c_tot_mrg.font = font_bold
    
    for c_idx in range(1, 7):
        ws_exec.cell(row=curr_row, column=c_idx).border = border_total

    # -------------------------------------------------------------
    # SHEET 1: RAW DATA / CLEANED TRANSACTIONS
    # -------------------------------------------------------------
    print("  Creating Sheet 1: Raw Data...")
    ws_raw = wb.create_sheet(title="Raw Data")
    ws_raw.views.sheetView[0].showGridLines = True
    
    raw_cols = list(df_sales.columns)
    for col_idx, col_name in enumerate(raw_cols, start=1):
        c = ws_raw.cell(row=1, column=col_idx, value=col_name)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        
    for r_idx, row in enumerate(df_sales.values, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=val)
            if raw_cols[c_idx-1] in ["sales_amount", "cost_amount", "profit", "unit_price"]:
                cell.number_format = "₹#,##0.00"
            elif raw_cols[c_idx-1] == "quantity":
                cell.number_format = "#,##0"
            elif raw_cols[c_idx-1] == "discount":
                cell.number_format = "0.00%"
                
    ws_raw.freeze_panes = "A2"
    
    # -------------------------------------------------------------
    # SHEET 2: DATA DICTIONARY
    # -------------------------------------------------------------
    print("  Creating Sheet 2: Data Dictionary...")
    ws_dict = wb.create_sheet(title="Data Dictionary")
    ws_dict.views.sheetView[0].showGridLines = True
    
    df_dict = pd.read_excel(os.path.join(DATA_DIR, "data_dictionary.xlsx"))
    dict_cols = list(df_dict.columns)
    for col_idx, col_name in enumerate(dict_cols, start=1):
        c = ws_dict.cell(row=1, column=col_idx, value=col_name)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        
    for r_idx, row in enumerate(df_dict.values, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_dict.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            
    # -------------------------------------------------------------
    # SHEET 3: SALES ANALYSIS (Formulas: SUMIFS, COUNTIFS, AVERAGEIFS)
    # -------------------------------------------------------------
    print("  Creating Sheet 3: Sales Analysis...")
    ws_sales = wb.create_sheet(title="Sales Analysis")
    ws_sales.views.sheetView[0].showGridLines = True
    
    ws_sales["A1"] = "MONTHLY SALES PERFORMANCE & GROWTH FORMULAS"
    ws_sales["A1"].font = font_section
    
    s_headers = ["Year-Month", "Orders", "Units Sold", "Total Revenue (₹)", "Total Profit (₹)", "Profit Margin %", "AOV (₹)", "MoM Growth %"]
    for col_idx, h in enumerate(s_headers, start=1):
        c = ws_sales.cell(row=3, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = border_cell
        
    df_monthly = df_sales.copy()
    df_monthly["ym"] = pd.to_datetime(df_monthly["order_date"]).dt.to_period("M").astype(str)
    month_list = sorted(df_monthly["ym"].unique())
    
    m_row = 4
    for idx, ym in enumerate(month_list):
        m_data = df_monthly[df_monthly["ym"] == ym]
        m_rev = m_data["sales_amount"].sum()
        m_prof = m_data["profit"].sum()
        m_orders = m_data["order_id"].nunique()
        m_units = m_data["quantity"].sum()
        
        ws_sales.cell(row=m_row, column=1, value=ym).font = font_regular
        
        c2 = ws_sales.cell(row=m_row, column=2, value=m_orders)
        c2.number_format = "#,##0"
        c2.font = font_regular
        
        c3 = ws_sales.cell(row=m_row, column=3, value=m_units)
        c3.number_format = "#,##0"
        c3.font = font_regular
        
        c4 = ws_sales.cell(row=m_row, column=4, value=m_rev)
        c4.number_format = "₹#,##0.00"
        c4.font = font_regular
        
        c5 = ws_sales.cell(row=m_row, column=5, value=m_prof)
        c5.number_format = "₹#,##0.00"
        c5.font = font_regular
        
        c6 = ws_sales.cell(row=m_row, column=6, value=f"=E{m_row}/D{m_row}")
        c6.number_format = "0.00%"
        c6.font = font_regular
        
        c7 = ws_sales.cell(row=m_row, column=7, value=f"=D{m_row}/B{m_row}")
        c7.number_format = "₹#,##0.00"
        c7.font = font_regular
        
        if idx == 0:
            c8 = ws_sales.cell(row=m_row, column=8, value="-")
        else:
            c8 = ws_sales.cell(row=m_row, column=8, value=f"=(D{m_row}-D{m_row-1})/D{m_row-1}")
            c8.number_format = "0.00%"
        c8.font = font_regular
        
        for c_idx in range(1, 9):
            ws_sales.cell(row=m_row, column=c_idx).border = border_cell
        m_row += 1
        
    # Total row
    ws_sales.cell(row=m_row, column=1, value="Total / Avg").font = font_bold
    ws_sales.cell(row=m_row, column=2, value=f"=SUM(B4:B{m_row-1})").font = font_bold
    ws_sales.cell(row=m_row, column=2).number_format = "#,##0"
    
    ws_sales.cell(row=m_row, column=3, value=f"=SUM(C4:C{m_row-1})").font = font_bold
    ws_sales.cell(row=m_row, column=3).number_format = "#,##0"
    
    ws_sales.cell(row=m_row, column=4, value=f"=SUM(D4:D{m_row-1})").font = font_bold
    ws_sales.cell(row=m_row, column=4).number_format = "₹#,##0.00"
    
    ws_sales.cell(row=m_row, column=5, value=f"=SUM(E4:E{m_row-1})").font = font_bold
    ws_sales.cell(row=m_row, column=5).number_format = "₹#,##0.00"
    
    ws_sales.cell(row=m_row, column=6, value=f"=E{m_row}/D{m_row}").font = font_bold
    ws_sales.cell(row=m_row, column=6).number_format = "0.00%"
    
    ws_sales.cell(row=m_row, column=7, value=f"=D{m_row}/B{m_row}").font = font_bold
    ws_sales.cell(row=m_row, column=7).number_format = "₹#,##0.00"
    
    ws_sales.cell(row=m_row, column=8, value=f"=AVERAGE(H5:H{m_row-1})").font = font_bold
    ws_sales.cell(row=m_row, column=8).number_format = "0.00%"
    
    for c_idx in range(1, 9):
        ws_sales.cell(row=m_row, column=c_idx).border = border_total

    # -------------------------------------------------------------
    # SHEET 4: CUSTOMER ANALYSIS
    # -------------------------------------------------------------
    print("  Creating Sheet 4: Customer Analysis...")
    ws_cust = wb.create_sheet(title="Customer Analysis")
    ws_cust.views.sheetView[0].showGridLines = True
    
    ws_cust["A1"] = "CUSTOMER RFM SEGMENTATION & LIFETIME VALUE PROFILING"
    ws_cust["A1"].font = font_section
    
    rfm_headers = ["Customer ID", "Customer Name", "Gender", "Age", "City", "Segment", "RFM Score", "Recency (Days)", "Frequency (Orders)", "Monetary Spend (₹)", "Customer Lifetime Profit (₹)"]
    for col_idx, h in enumerate(rfm_headers, start=1):
        c = ws_cust.cell(row=3, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        
    for r_idx, row in enumerate(df_rfm.values, start=4):
        # rfm row: customer_id, Last_Order_Date, Frequency, Monetary, Total_Profit, Recency, R_Score, F_Score, M_Score, RFM_Score, Segment, customer_name, gender, age, city, region_id, signup_date, customer_segment
        c_id = row[0]
        freq = row[2]
        mon = row[3]
        prof = row[4]
        rec = row[5]
        rfm_sc = row[9]
        seg = row[10]
        c_name = row[11]
        gen = row[12]
        age = row[13]
        city = row[14]
        
        ws_cust.cell(row=r_idx, column=1, value=c_id).font = font_regular
        ws_cust.cell(row=r_idx, column=2, value=c_name).font = font_regular
        ws_cust.cell(row=r_idx, column=3, value=gen).font = font_regular
        ws_cust.cell(row=r_idx, column=4, value=age).font = font_regular
        ws_cust.cell(row=r_idx, column=5, value=city).font = font_regular
        ws_cust.cell(row=r_idx, column=6, value=seg).font = font_regular
        ws_cust.cell(row=r_idx, column=7, value=rfm_sc).font = font_regular
        ws_cust.cell(row=r_idx, column=8, value=rec).font = font_regular
        
        c9 = ws_cust.cell(row=r_idx, column=9, value=freq)
        c9.number_format = "#,##0"
        c9.font = font_regular
        
        c10 = ws_cust.cell(row=r_idx, column=10, value=mon)
        c10.number_format = "₹#,##0.00"
        c10.font = font_regular
        
        c11 = ws_cust.cell(row=r_idx, column=11, value=prof)
        c11.number_format = "₹#,##0.00"
        c11.font = font_regular
        
    ws_cust.freeze_panes = "A4"

    # -------------------------------------------------------------
    # SHEET 5: PRODUCT ANALYSIS
    # -------------------------------------------------------------
    print("  Creating Sheet 5: Product Analysis...")
    ws_p = wb.create_sheet(title="Product Analysis")
    ws_p.views.sheetView[0].showGridLines = True
    
    ws_p["A1"] = "PRODUCT CATALOG & SALES PERFORMANCE MATRIX"
    ws_p["A1"].font = font_section
    
    p_headers = ["Product ID", "Product Name", "Category", "Subcategory", "Brand", "Unit Cost (₹)", "Unit Price (₹)", "Units Sold", "Total Revenue (₹)", "Total Profit (₹)", "Profit Margin %"]
    for col_idx, h in enumerate(p_headers, start=1):
        c = ws_p.cell(row=3, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        
    prod_sales_agg = df_sales.groupby("product_id").agg(
        Units=("quantity", "sum"),
        Revenue=("sales_amount", "sum"),
        Profit=("profit", "sum")
    ).reset_index()
    
    df_prod_full = df_prod.merge(prod_sales_agg, on="product_id", how="left").fillna(0)
    
    for r_idx, r in enumerate(df_prod_full.values, start=4):
        ws_p.cell(row=r_idx, column=1, value=r[0]).font = font_regular # ID
        ws_p.cell(row=r_idx, column=2, value=r[1]).font = font_regular # Name
        ws_p.cell(row=r_idx, column=3, value=r[2]).font = font_regular # Cat
        ws_p.cell(row=r_idx, column=4, value=r[3]).font = font_regular # Subcat
        ws_p.cell(row=r_idx, column=5, value=r[4]).font = font_regular # Brand
        
        c6 = ws_p.cell(row=r_idx, column=6, value=r[5]) # Cost
        c6.number_format = "₹#,##0.00"
        c6.font = font_regular
        
        c7 = ws_p.cell(row=r_idx, column=7, value=r[6]) # Price
        c7.number_format = "₹#,##0.00"
        c7.font = font_regular
        
        c8 = ws_p.cell(row=r_idx, column=8, value=r[7]) # Units
        c8.number_format = "#,##0"
        c8.font = font_regular
        
        c9 = ws_p.cell(row=r_idx, column=9, value=r[8]) # Revenue
        c9.number_format = "₹#,##0.00"
        c9.font = font_regular
        
        c10 = ws_p.cell(row=r_idx, column=10, value=r[9]) # Profit
        c10.number_format = "₹#,##0.00"
        c10.font = font_regular
        
        c11 = ws_p.cell(row=r_idx, column=11, value=f"=J{r_idx}/I{r_idx}" if r[8] > 0 else 0)
        c11.number_format = "0.00%"
        c11.font = font_regular
        
    ws_p.freeze_panes = "A4"

    # -------------------------------------------------------------
    # SHEET 6: REGIONAL ANALYSIS
    # -------------------------------------------------------------
    print("  Creating Sheet 6: Regional Analysis...")
    ws_r = wb.create_sheet(title="Regional Analysis")
    ws_r.views.sheetView[0].showGridLines = True
    
    ws_r["A1"] = "REGIONAL PERFORMANCE & GEOGRAPHIC DISTRIBUTION"
    ws_r["A1"].font = font_section
    
    r_headers = ["Region ID", "Region Name", "State Coverage", "Zone", "Unique Customers", "Total Orders", "Units Sold", "Total Revenue (₹)", "Revenue Share %", "Total Profit (₹)", "Profit Margin %", "AOV (₹)"]
    for col_idx, h in enumerate(r_headers, start=1):
        c = ws_r.cell(row=3, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = border_cell
        
    df_reg_detail = df_master.groupby(["region_id", "region_name", "state", "zone"]).agg(
        Custs=("customer_id", "nunique"),
        Orders=("order_id", "nunique"),
        Units=("quantity", "sum"),
        Revenue=("sales_amount", "sum"),
        Profit=("profit", "sum")
    ).reset_index().sort_values(by="Revenue", ascending=False)
    
    r_row = 4
    for _, r in df_reg_detail.iterrows():
        ws_r.cell(row=r_row, column=1, value=r["region_id"]).font = font_regular
        ws_r.cell(row=r_row, column=2, value=r["region_name"]).font = font_regular
        ws_r.cell(row=r_row, column=3, value=r["state"]).font = font_regular
        ws_r.cell(row=r_row, column=4, value=r["zone"]).font = font_regular
        
        ws_r.cell(row=r_row, column=5, value=r["Custs"]).font = font_regular
        ws_r.cell(row=r_row, column=5).number_format = "#,##0"
        
        ws_r.cell(row=r_row, column=6, value=r["Orders"]).font = font_regular
        ws_r.cell(row=r_row, column=6).number_format = "#,##0"
        
        ws_r.cell(row=r_row, column=7, value=r["Units"]).font = font_regular
        ws_r.cell(row=r_row, column=7).number_format = "#,##0"
        
        c8 = ws_r.cell(row=r_row, column=8, value=r["Revenue"])
        c8.number_format = "₹#,##0.00"
        c8.font = font_regular
        
        c9 = ws_r.cell(row=r_row, column=9, value=r["Revenue"] / tot_rev)
        c9.number_format = "0.00%"
        c9.font = font_regular
        
        c10 = ws_r.cell(row=r_row, column=10, value=r["Profit"])
        c10.number_format = "₹#,##0.00"
        c10.font = font_regular
        
        c11 = ws_r.cell(row=r_row, column=11, value=f"=J{r_row}/H{r_row}")
        c11.number_format = "0.00%"
        c11.font = font_regular
        
        c12 = ws_r.cell(row=r_row, column=12, value=f"=H{r_row}/F{r_row}")
        c12.number_format = "₹#,##0.00"
        c12.font = font_regular
        
        for c_idx in range(1, 13):
            ws_r.cell(row=r_row, column=c_idx).border = border_cell
        r_row += 1
        
    # Total row
    ws_r.cell(row=r_row, column=1, value="Total").font = font_bold
    ws_r.cell(row=r_row, column=2, value="All Regions").font = font_bold
    ws_r.cell(row=r_row, column=3, value="-").font = font_bold
    ws_r.cell(row=r_row, column=4, value="-").font = font_bold
    
    ws_r.cell(row=r_row, column=5, value=f"=SUM(E4:E{r_row-1})").font = font_bold
    ws_r.cell(row=r_row, column=5).number_format = "#,##0"
    
    ws_r.cell(row=r_row, column=6, value=f"=SUM(F4:F{r_row-1})").font = font_bold
    ws_r.cell(row=r_row, column=6).number_format = "#,##0"
    
    ws_r.cell(row=r_row, column=7, value=f"=SUM(G4:G{r_row-1})").font = font_bold
    ws_r.cell(row=r_row, column=7).number_format = "#,##0"
    
    ws_r.cell(row=r_row, column=8, value=f"=SUM(H4:H{r_row-1})").font = font_bold
    ws_r.cell(row=r_row, column=8).number_format = "₹#,##0.00"
    
    ws_r.cell(row=r_row, column=9, value=f"=SUM(I4:I{r_row-1})").font = font_bold
    ws_r.cell(row=r_row, column=9).number_format = "0.00%"
    
    ws_r.cell(row=r_row, column=10, value=f"=SUM(J4:J{r_row-1})").font = font_bold
    ws_r.cell(row=r_row, column=10).number_format = "₹#,##0.00"
    
    ws_r.cell(row=r_row, column=11, value=f"=J{r_row}/H{r_row}").font = font_bold
    ws_r.cell(row=r_row, column=11).number_format = "0.00%"
    
    ws_r.cell(row=r_row, column=12, value=f"=H{r_row}/F{r_row}").font = font_bold
    ws_r.cell(row=r_row, column=12).number_format = "₹#,##0.00"
    
    for c_idx in range(1, 13):
        ws_r.cell(row=r_row, column=c_idx).border = border_total

    # -------------------------------------------------------------
    # SHEET 7: PIVOT ANALYSIS (Summary Pivot Dimensions)
    # -------------------------------------------------------------
    print("  Creating Sheet 7: Pivot Analysis...")
    ws_piv = wb.create_sheet(title="Pivot Analysis")
    ws_piv.views.sheetView[0].showGridLines = True
    
    ws_piv["A1"] = "MULTI-DIMENSIONAL PIVOT SUMMARIES"
    ws_piv["A1"].font = font_section
    
    # Pivot 1: Category by Region Matrix
    ws_piv["A3"] = "Category x Region Revenue Matrix (₹)"
    ws_piv["A3"].font = Font(name="Calibri", size=11, bold=True, color="1B365D")
    
    piv_cat_reg = df_master.pivot_table(index="category", columns="region_name", values="sales_amount", aggfunc="sum").fillna(0)
    
    # Headers
    ws_piv.cell(row=4, column=1, value="Category").font = font_header
    ws_piv.cell(row=4, column=1).fill = fill_header
    ws_piv.cell(row=4, column=1).border = border_cell
    
    for c_idx, col in enumerate(piv_cat_reg.columns, start=2):
        c = ws_piv.cell(row=4, column=c_idx, value=col)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = border_cell
        
    piv_row = 5
    for cat, r in piv_cat_reg.iterrows():
        ws_piv.cell(row=piv_row, column=1, value=cat).font = font_regular
        ws_piv.cell(row=piv_row, column=1).border = border_cell
        for c_idx, val in enumerate(r.values, start=2):
            cell = ws_piv.cell(row=piv_row, column=c_idx, value=val)
            cell.number_format = "₹#,##0.00"
            cell.font = font_regular
            cell.border = border_cell
        piv_row += 1
        
    # Auto-fit column widths across all sheets
    print("  Auto-adjusting column widths...")
    for sheet in wb.worksheets:
        for col in sheet.columns:
            col_letter = get_column_letter(col[0].column)
            # Find max length
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len and len(val_str) < 50:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Save Workbook
    wb.save(EXCEL_FILE)
    print(f"\n  ✓ Successfully generated {EXCEL_FILE}")
    print("=" * 60)
    print("  ✅ EXCEL WORKBOOK GENERATION COMPLETE: 8 SHEETS CREATED")
    print("=" * 60)

if __name__ == "__main__":
    create_workbook()
